import pandas as pd
import numpy as np
import random
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class DataManager:
    def __init__(self):
        self.question_bank = None
        self.results_file = "考核结果.xlsx"
        self.current_student = None
        self.current_questions = []
        
    def load_question_bank(self, file_path):
        """加载题库Excel文件"""
        try:
            self.question_bank = pd.read_excel(file_path)
            # 假设题库只有一列问题，列名为'问题'或自动检测
            if len(self.question_bank.columns) == 1:
                self.question_bank.columns = ['问题']
            return True, f"成功加载题库，共 {len(self.question_bank)} 道题目"
        except Exception as e:
            return False, f"加载题库失败: {str(e)}"
    
    def get_random_questions(self, num_questions):
        """随机抽取指定数量的题目"""
        if self.question_bank is None:
            return []
        
        if num_questions > len(self.question_bank):
            num_questions = len(self.question_bank)
            
        # 使用随机抽样，不重复
        random_indices = random.sample(range(len(self.question_bank)), num_questions)
        selected_questions = []
        
        for idx in random_indices:
            question = {
                '序号': idx + 1,
                '问题': self.question_bank.iloc[idx]['问题'],
                '知识点分数': 0,
                '流利程度': 0,
                '知识扩展': 0,
                '思路严谨': 0,
                '总分': 0,
                '评语': ''
            }
            selected_questions.append(question)
        
        self.current_questions = selected_questions
        return selected_questions
    
    def save_results(self, student_name, questions, statistics):
        """保存考核结果到Excel"""
        try:
            # 准备数据
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            results_df = pd.DataFrame(questions)
            
            # 添加统计信息作为新行
            stats_row = {
                '序号': '统计',
                '问题': '平均分',
                '知识点分数': statistics['avg_knowledge'],
                '流利程度': statistics['avg_fluency'],
                '知识扩展': statistics['avg_expansion'],
                '思路严谨': statistics['avg_rigor'],
                '总分': statistics['avg_total'],
                '评语': f"考核时间: {current_time}"
            }
            results_df = pd.concat([results_df, pd.DataFrame([stats_row])], ignore_index=True)
            
            # 保存到Excel
            if os.path.exists(self.results_file):
                with pd.ExcelWriter(self.results_file, mode='a', engine='openpyxl', 
                                  if_sheet_exists='replace') as writer:
                    # 移除可能存在的同名sheet（如果已存在）
                    workbook = writer.book
                    if student_name in workbook.sheetnames:
                        std = workbook[student_name]
                        workbook.remove(std)
                    
                    results_df.to_excel(writer, sheet_name=student_name, index=False)
            else:
                results_df.to_excel(self.results_file, sheet_name=student_name, index=False)
            
            return True, "考核结果已保存"
            
        except Exception as e:
            return False, f"保存失败: {str(e)}"
    
    def get_student_sheets(self):
        """获取已有学生的sheet列表"""
        if not os.path.exists(self.results_file):
            return []
        
        try:
            workbook = load_workbook(self.results_file, read_only=True)
            return workbook.sheetnames
        except:
            return []
    
    def load_student_history(self, student_name):
        """加载学生历史考核记录"""
        try:
            df = pd.read_excel(self.results_file, sheet_name=student_name)
            return df
        except:
            return pd.DataFrame()

class ScoringSystem:
    @staticmethod
    def calculate_statistics(questions):
        """计算各项平均分"""
        if not questions:
            return {
                'avg_knowledge': 0,
                'avg_fluency': 0,
                'avg_expansion': 0,
                'avg_rigor': 0,
                'avg_total': 0
            }
        
        knowledge_scores = [q['知识点分数'] for q in questions]
        fluency_scores = [q['流利程度'] for q in questions]
        expansion_scores = [q['知识扩展'] for q in questions]
        rigor_scores = [q['思路严谨'] for q in questions]
        total_scores = [q['总分'] for q in questions]
        
        return {
            'avg_knowledge': np.mean(knowledge_scores),
            'avg_fluency': np.mean(fluency_scores),
            'avg_expansion': np.mean(expansion_scores),
            'avg_rigor': np.mean(rigor_scores),
            'avg_total': np.mean(total_scores)
        }
    
    @staticmethod
    def calculate_total_score(question):
        """计算单题总分"""
        scores = [
            question['知识点分数'],
            question['流利程度'],
            question['知识扩展'],
            question['思路严谨']
        ]
        return min(100, sum(scores))  # 确保不超过100分